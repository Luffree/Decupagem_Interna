"""
Excel generator for decupagem output.
Produces a formatted .xlsx with:
  1. Resumo
  2. Itens Detalhados  (28 colunas A-AB, idêntico à referência)
  3. Por Seção         (linhas por item, agrupadas por seção)
  4. Medições          (uma linha por dimensão)
  5. Pendências        (colunas alinhadas à referência)
  6. Legenda e Status
  7+. Uma aba por fornecedor  (ao final, ordenadas)
"""
from __future__ import annotations

import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Palette ───────────────────────────────────────────────────────────────────
_NAV   = "1F3864"   # cabeçalho principal (azul marinho)
_BLU   = "2E75B6"   # sub-cabeçalho
_ALT   = "DCE6F1"   # linha alternada
_ALTO  = "FF6B6B"   # nível alto
_MEDIO = "FFD93D"   # nível médio
_BAIXO = "6BCB77"   # nível baixo
_CRIT  = "C0392B"   # crítico
_RES   = "E8F0FE"   # rótulo resumo
_WHITE = "FFFFFF"
_GRAY  = "F2F2F2"
_BORD  = "BDC3C7"

NIVEL_FILL = {
    "alto":   _ALTO,
    "médio":  _MEDIO,
    "medio":  _MEDIO,
    "baixo":  _BAIXO,
    "crítico":_CRIT,
    "critico":_CRIT,
}

NIVEL_ROW_FILL = {
    "alto":   "FFF0F0",
    "crítico":"FFE0E0",
    "critico":"FFE0E0",
}


def _solid(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _border(color=_BORD, style="thin") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold=False, size=9, color="000000", name="Calibri") -> Font:
    return Font(name=name, bold=bold, size=size, color=color)


def _align(h="left", v="top", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _hdr_cell(ws, row: int, col: int, value: str,
              bg=_NAV, fg=_WHITE, size=10, bold=True, h="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Calibri", bold=bold, size=size, color=fg)
    cell.fill      = _solid(bg)
    cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
    cell.border    = _border()
    return cell


def _data_cell(ws, row: int, col: int, value,
               bg=None, h="left", bold=False, size=9):
    cell = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    cell.font      = _font(bold=bold, size=size)
    cell.alignment = _align(h=h)
    cell.border    = _border()
    if bg:
        cell.fill = _solid(bg)
    return cell


def _set_widths(ws, widths: dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def _autofilter(ws, row=1):
    ws.auto_filter.ref = ws.dimensions


def _row_height(ws, row: int, h: float):
    ws.row_dimensions[row].height = h


# ─────────────────────────────────────────────────────────────────────────────
# 1. RESUMO
# ─────────────────────────────────────────────────────────────────────────────

def _build_resumo(ws, data: dict, filename: str):
    ws.sheet_view.showGridLines = False

    resumo = data.get("resumo", {})
    n_med  = len(data.get("medicoes", []))
    n_pend = len(data.get("pendencias", []))

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "DECUPAGEM TÉCNICA — RESUMO DO PROJETO"
    t.font  = Font(name="Calibri", bold=True, size=14, color=_WHITE)
    t.fill  = _solid(_NAV)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 32)

    ws.merge_cells("D1:F1")
    t2 = ws["D1"]
    t2.value = "DISTRIBUIÇÃO DE ITENS"
    t2.font  = Font(name="Calibri", bold=True, size=11, color=_WHITE)
    t2.fill  = _solid(_NAV)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    # ── Dados principais (A–B) ────────────────────────────────────────────────
    fields = [
        ("Campo", "Valor"),  # header
        ("Cliente",                    resumo.get("cliente", "Não informado")),
        ("Projeto",                    resumo.get("projeto", "Não informado")),
        ("Número total de páginas/slides analisados",
                                       resumo.get("total_paginas_slides", "Não informado")),
        ("Quantidade total de itens",  resumo.get("total_itens", 0)),
        ("Quantidade total de medições", n_med),
        ("Quantidade total de pendências/ações", n_pend),
        ("Observações gerais",         resumo.get("observacoes_gerais", "")),
    ]

    for r_idx, (campo, valor) in enumerate(fields, start=1):
        row = r_idx + 1
        is_header = r_idx == 1
        bg_a = _BLU if is_header else _RES
        fg_a = _WHITE if is_header else _NAV

        ca = ws.cell(row=row, column=1, value=campo)
        ca.font      = Font(name="Calibri", bold=True, size=9 if not is_header else 10, color=fg_a)
        ca.fill      = _solid(bg_a)
        ca.alignment = _align(h="left" if not is_header else "center", v="center")
        ca.border    = _border()

        cb = ws.cell(row=row, column=2, value=str(valor))
        cb.font      = _font(size=9)
        cb.alignment = _align(h="left", v="center")
        cb.border    = _border()
        _row_height(ws, row, 18 if campo != "Observações gerais" else 40)

    # ── Resumo D–F (Tipo / Chave / Quantidade) ────────────────────────────────
    # Header row 2
    for col, label in [(4, "Tipo de resumo"), (5, "Chave"), (6, "Quantidade")]:
        _hdr_cell(ws, 2, col, label, bg=_BLU, size=9)

    right_rows = []
    for s in resumo.get("secoes", []):
        right_rows.append(("Quantidade por seção", s.get("nome", ""), s.get("quantidade_itens", 0)))
    for f in resumo.get("fornecedores", []):
        right_rows.append(("Quantidade por fornecedor", f.get("nome", ""), f.get("quantidade_itens", 0)))
    for st in resumo.get("status_summary", []):
        right_rows.append(("Quantidade por status de compra/locação",
                           st.get("status", ""), st.get("quantidade", 0)))

    for i, (tipo, chave, qtd) in enumerate(right_rows):
        row = i + 3
        bg  = _GRAY if i % 2 == 0 else None
        _data_cell(ws, row, 4, tipo,  bg=bg, h="left")
        _data_cell(ws, row, 5, chave, bg=bg, h="left")
        _data_cell(ws, row, 6, qtd,   bg=bg, h="center")
        _row_height(ws, row, 16)

    _set_widths(ws, {1: 36, 2: 55, 3: 3, 4: 32, 5: 38, 6: 14})


# ─────────────────────────────────────────────────────────────────────────────
# 2. ITENS DETALHADOS  (A-AB, 28 colunas — idêntico à referência)
# ─────────────────────────────────────────────────────────────────────────────

ITENS_HEADERS = [
    "ID", "Página/Slide de origem", "Seção", "Sub-seção / Ambiente",
    "Fornecedor", "Categoria", "Item", "Descrição detalhada",
    "Quantidade", "Unidade", "Medida original",
    "Largura", "Altura", "Profundidade", "Comprimento", "Espessura",
    "Área", "Metro linear",
    "Material", "Acabamento / Cor", "Tipo de produção",
    "Status da arte", "Status da compra/locação",
    "Responsável / área", "Observações técnicas",
    "Pendência / ação necessária", "Nível de atenção", "Fonte da informação",
]

ITENS_FIELDS = [
    "id", "pagina_slide_origem", "secao", "sub_secao_ambiente",
    "fornecedor", "categoria", "item", "descricao_detalhada",
    "quantidade", "unidade", "medida_original",
    "largura", "altura", "profundidade", "comprimento", "espessura",
    "area", "metro_linear",
    "material", "acabamento_cor", "tipo_producao",
    "status_arte", "status_compra_locacao",
    "responsavel_area", "observacoes_tecnicas",
    "pendencia_acao_necessaria", "nivel_atencao", "fonte_informacao",
]

# Column widths matching reference (A=10 … AB=42)
ITENS_WIDTHS = {
    1:10,  2:28,  3:22,  4:26,  5:24,  6:24,  7:35,  8:48,
    9:12, 10:13, 11:24, 12:11, 13:11, 14:14, 15:14, 16:12,
   17:12, 18:14, 19:24, 20:24, 21:24, 22:24, 23:28, 24:24,
   25:42, 26:42, 27:16, 28:42,
}

_CENTER_COLS_ITENS = {1, 2, 9, 10, 12, 13, 14, 15, 16, 17, 18, 22, 23, 27}


def _build_itens(ws, itens: list[dict]):
    # Header row
    for col, label in enumerate(ITENS_HEADERS, start=1):
        _hdr_cell(ws, 1, col, label, size=9)
    _row_height(ws, 1, 26)
    _freeze(ws, "A2")
    _set_widths(ws, ITENS_WIDTHS)

    for r_idx, item in enumerate(itens, start=2):
        nivel    = str(item.get("nivel_atencao", "")).strip().lower()
        row_bg   = NIVEL_ROW_FILL.get(nivel)
        nivel_bg = NIVEL_FILL.get(nivel)

        for c_idx, field in enumerate(ITENS_FIELDS, start=1):
            val   = item.get(field, "") or ""
            h_al  = "center" if c_idx in _CENTER_COLS_ITENS else "left"
            bg    = row_bg if (row_bg and nivel in ("alto", "crítico", "critico")) else \
                    (_ALT if r_idx % 2 == 0 else None)
            _data_cell(ws, r_idx, c_idx, val, bg=bg, h=h_al)

        # Colour the nivel_atencao cell itself
        if nivel_bg:
            cell = ws.cell(row=r_idx, column=27)
            cell.fill = _solid(nivel_bg)
            cell.font = Font(name="Calibri", bold=True, size=9)

        _row_height(ws, r_idx, 38)

    _autofilter(ws)


# ─────────────────────────────────────────────────────────────────────────────
# 3. POR SEÇÃO  (linhas individuais por item, agrupadas por seção)
# ─────────────────────────────────────────────────────────────────────────────

SECAO_HEADERS = [
    "Seção", "Sub-seção / Ambiente", "ID", "Item",
    "Categoria", "Fornecedor", "Quantidade", "Unidade",
    "Status da compra/locação", "Nível de atenção", "Pendência / ação necessária",
]
SECAO_WIDTHS = {1:26, 2:28, 3:10, 4:40, 5:26, 6:26, 7:10, 8:10, 9:30, 10:14, 11:45}


def _build_por_secao(ws, itens: list[dict]):
    for col, label in enumerate(SECAO_HEADERS, start=1):
        _hdr_cell(ws, 1, col, label, size=9)
    _row_height(ws, 1, 24)
    _freeze(ws, "A2")
    _set_widths(ws, SECAO_WIDTHS)

    # Sort by seção then sub-seção
    sorted_itens = sorted(itens, key=lambda x: (
        x.get("secao", ""), x.get("sub_secao_ambiente", "")
    ))

    prev_sec = None
    for r_idx, item in enumerate(sorted_itens, start=2):
        sec      = item.get("secao", "")
        sub      = item.get("sub_secao_ambiente", "")
        nivel    = str(item.get("nivel_atencao", "")).strip().lower()
        nivel_bg = NIVEL_FILL.get(nivel)

        # Section header stripe
        is_new_sec = sec != prev_sec
        row_bg     = _ALT if r_idx % 2 == 0 else None

        values = [
            sec,
            sub,
            item.get("id", ""),
            item.get("item", ""),
            item.get("categoria", ""),
            item.get("fornecedor", ""),
            item.get("quantidade", ""),
            item.get("unidade", ""),
            item.get("status_compra_locacao", ""),
            item.get("nivel_atencao", ""),
            item.get("pendencia_acao_necessaria", ""),
        ]
        _CENTER = {3, 7, 8, 10}
        for c_idx, val in enumerate(values, start=1):
            h_al = "center" if c_idx in _CENTER else "left"
            bold = c_idx == 1 and is_new_sec
            _data_cell(ws, r_idx, c_idx, val, bg=row_bg, h=h_al, bold=bold)

        if nivel_bg:
            cell = ws.cell(row=r_idx, column=10)
            cell.fill = _solid(nivel_bg)
            cell.font = Font(name="Calibri", bold=True, size=9)

        if is_new_sec:
            for c in range(1, len(SECAO_HEADERS) + 1):
                ws.cell(row=r_idx, column=c).fill = _solid("D6E4F0")
                ws.cell(row=r_idx, column=c).font = Font(name="Calibri", bold=True, size=9)

        _row_height(ws, r_idx, 28)
        prev_sec = sec

    _autofilter(ws)


# ─────────────────────────────────────────────────────────────────────────────
# 4. MEDIÇÕES  (uma linha por dimensão)
# ─────────────────────────────────────────────────────────────────────────────

MED_HEADERS = [
    "ID", "Item", "Página",
    "Medida original", "Unidade original",
    "Medida convertida para metros",
    "Tipo de medida",
    "Observação sobre confiabilidade da leitura",
    "Fonte da informação",
]
MED_WIDTHS = {1:10, 2:38, 3:20, 4:22, 5:14, 6:18, 7:20, 8:44, 9:40}

_CONF_FILL = {"alta": _BAIXO, "alta - medida textual no arquivo": _BAIXO,
              "média": _MEDIO, "media": _MEDIO, "baixa": _ALTO}


def _expand_medicoes(itens: list[dict]) -> list[dict]:
    """Expand each item with measurements into one row per dimension."""
    rows = []
    dim_fields = [
        ("largura",      "largura"),
        ("altura",       "altura"),
        ("profundidade", "profundidade"),
        ("comprimento",  "comprimento"),
        ("espessura",    "espessura"),
        ("area",         "área"),
        ("metro_linear", "metro linear"),
    ]
    for item in itens:
        med_orig = (item.get("medida_original") or "").strip()
        if not med_orig:
            continue
        fonte = item.get("fonte_informacao", "")
        page  = item.get("pagina_slide_origem", "")

        # Unit from medida_original
        u_m = re.search(r'(cm|mm|m\b)', med_orig, re.IGNORECASE)
        unit = u_m.group(1) if u_m else ""

        added = False
        for field, tipo in dim_fields:
            val = (item.get(field) or "").strip()
            if not val:
                continue
            # Convert to meters
            try:
                v_f  = float(val.replace(",", "."))
                u_low = unit.lower()
                if u_low == "cm":
                    metros = f"{v_f/100:.3f}"
                elif u_low == "mm":
                    metros = f"{v_f/1000:.4f}"
                else:
                    metros = f"{v_f:.3f}"
            except ValueError:
                metros = "A confirmar"

            rows.append({
                "id_item":      item.get("id", ""),
                "item":         item.get("item", ""),
                "pagina":       page,
                "medida_orig":  med_orig,
                "unidade_orig": unit,
                "metros":       metros,
                "tipo":         tipo,
                "confiab":      "Alta - medida textual no arquivo",
                "fonte":        fonte,
            })
            added = True

        # If no individual dimensions, emit one raw row
        if not added:
            rows.append({
                "id_item":      item.get("id", ""),
                "item":         item.get("item", ""),
                "pagina":       page,
                "medida_orig":  med_orig,
                "unidade_orig": unit,
                "metros":       "A confirmar",
                "tipo":         "medida",
                "confiab":      "Média",
                "fonte":        fonte,
            })
    return rows


def _build_medicoes(ws, itens: list[dict]):
    for col, label in enumerate(MED_HEADERS, start=1):
        _hdr_cell(ws, 1, col, label, size=9)
    _row_height(ws, 1, 24)
    _freeze(ws, "A2")
    _set_widths(ws, MED_WIDTHS)

    rows = _expand_medicoes(itens)

    for r_idx, row in enumerate(rows, start=2):
        conf_key = row["confiab"].lower()
        conf_bg  = _CONF_FILL.get(conf_key) or _CONF_FILL.get("média")
        alt_bg   = _ALT if r_idx % 2 == 0 else None

        values = [
            row["id_item"], row["item"], row["pagina"],
            row["medida_orig"], row["unidade_orig"],
            row["metros"], row["tipo"], row["confiab"], row["fonte"],
        ]
        _CENTER = {1, 3, 5, 6, 7}
        for c_idx, val in enumerate(values, start=1):
            h = "center" if c_idx in _CENTER else "left"
            _data_cell(ws, r_idx, c_idx, val, bg=alt_bg, h=h)

        # Colour confiabilidade cell
        ws.cell(row=r_idx, column=8).fill = _solid(conf_bg)
        ws.cell(row=r_idx, column=8).font = Font(name="Calibri", bold=True, size=9)

        _row_height(ws, r_idx, 20)

    _autofilter(ws)


# ─────────────────────────────────────────────────────────────────────────────
# 5. PENDÊNCIAS
# ─────────────────────────────────────────────────────────────────────────────

PEND_HEADERS = [
    "ID", "Página/Slide", "Seção", "Sub-seção / Ambiente", "Item",
    "Pendência / ação necessária",
    "Status da arte", "Status da compra/locação",
    "Nível de atenção", "Fonte da informação",
]
PEND_WIDTHS = {1:10, 2:24, 3:24, 4:26, 5:38, 6:50, 7:24, 8:28, 9:14, 10:40}


def _build_pendencias(ws, pendencias: list[dict], item_lookup: dict):
    for col, label in enumerate(PEND_HEADERS, start=1):
        _hdr_cell(ws, 1, col, label, size=9)
    _row_height(ws, 1, 24)
    _freeze(ws, "A2")
    _set_widths(ws, PEND_WIDTHS)

    for r_idx, pend in enumerate(pendencias, start=2):
        item_id  = pend.get("id_item", "")
        item_obj = item_lookup.get(item_id, {})
        nivel    = str(pend.get("nivel_urgencia", "")).strip().lower()
        nivel_bg = NIVEL_FILL.get(nivel)
        alt_bg   = _ALT if r_idx % 2 == 0 else None

        values = [
            item_id,
            pend.get("pagina_slide", ""),
            pend.get("secao", ""),
            item_obj.get("sub_secao_ambiente", ""),
            pend.get("item", ""),
            pend.get("descricao_pendencia", ""),
            item_obj.get("status_arte", ""),
            item_obj.get("status_compra_locacao", ""),
            pend.get("nivel_urgencia", ""),
            item_obj.get("fonte_informacao", ""),
        ]
        _CENTER = {1, 2, 9}
        for c_idx, val in enumerate(values, start=1):
            h = "center" if c_idx in _CENTER else "left"
            _data_cell(ws, r_idx, c_idx, val,
                       bg=("FFF5F5" if nivel in ("crítico","critico","alto") else alt_bg),
                       h=h)

        if nivel_bg:
            cell = ws.cell(row=r_idx, column=9)
            cell.fill = _solid(nivel_bg)
            cell.font = Font(name="Calibri", bold=True, size=9,
                             color=_WHITE if nivel in ("crítico","critico") else "000000")

        _row_height(ws, r_idx, 30)

    _autofilter(ws)


# ─────────────────────────────────────────────────────────────────────────────
# 6. LEGENDA / STATUS
# ─────────────────────────────────────────────────────────────────────────────

def _build_legenda(ws, legenda: list[dict]):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "LEGENDA DE CORES E STATUS"
    t.font  = Font(name="Calibri", bold=True, size=13, color=_WHITE)
    t.fill  = _solid(_NAV)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 30)

    for col, label in enumerate(["Cor", "Significado", "Fornecedor", "Status relacionado", "Observações"], start=1):
        _hdr_cell(ws, 2, col, label, bg=_BLU, size=9)

    # Custom legenda from file
    for r_idx, leg in enumerate(legenda, start=3):
        hex_c = (leg.get("cor_hex") or "").replace("#", "").strip()
        _data_cell(ws, r_idx, 1, f"#{hex_c}" if hex_c else leg.get("cor_hex", "Não informado"))
        sample = ws.cell(row=r_idx, column=2, value="")
        if hex_c and len(hex_c) == 6:
            try:
                sample.fill = _solid(hex_c)
            except Exception:
                pass
        sample.border = _border()
        for c_idx, key in enumerate(["significado", "fornecedor", "status_relacionado", "observacoes"], start=3):
            _data_cell(ws, r_idx, c_idx, leg.get(key, ""))
        _row_height(ws, r_idx, 20)

    # Built-in legend
    sep_row = len(legenda) + 4
    ws.merge_cells(f"A{sep_row}:E{sep_row}")
    t2 = ws[f"A{sep_row}"]
    t2.value = "LEGENDA INTERNA — NÍVEL DE ATENÇÃO"
    t2.font  = Font(name="Calibri", bold=True, size=10, color=_WHITE)
    t2.fill  = _solid(_BLU)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, sep_row, 22)

    built_in = [
        ("Verde claro",         _BAIXO, "Baixo",           "Todos", "Sem alerta relevante."),
        ("Amarelo claro",       _MEDIO, "Médio",            "Todos", "Requer ação antes da compra/produção."),
        ("Vermelho claro",      _ALTO,  "Alto",             "Todos", "Verificar, definir, conferir ou medida aproximada."),
        ("Vermelho escuro",     _CRIT,  "Crítico",          "Todos", "Pendência bloqueante ou urgente."),
        ("Azul claro",          "9DC3E6","Locação necessária","Fornecedor de locação","Item indica locação."),
    ]
    for i, (cor_nome, hex_c, nivel, forn, obs) in enumerate(built_in):
        row = sep_row + 1 + i
        _data_cell(ws, row, 1, cor_nome)
        ws.cell(row=row, column=2).fill   = _solid(hex_c)
        ws.cell(row=row, column=2).border = _border()
        _data_cell(ws, row, 3, nivel, bold=True)
        _data_cell(ws, row, 4, forn)
        _data_cell(ws, row, 5, obs)
        _row_height(ws, row, 20)

    _set_widths(ws, {1:16, 2:10, 3:32, 4:26, 5:46})


# ─────────────────────────────────────────────────────────────────────────────
# 7. ABA POR FORNECEDOR
# ─────────────────────────────────────────────────────────────────────────────

FORN_HEADERS = [
    "ID", "Seção", "Sub-seção / Ambiente", "Item", "Descrição detalhada",
    "Quantidade", "Unidade",
    "Medida original", "Largura", "Altura", "Profundidade",
    "Material", "Acabamento / Cor", "Tipo de produção",
    "Status da arte", "Status da compra/locação",
    "Observações técnicas", "Pendência / ação necessária", "Nível de atenção",
]
FORN_FIELDS = [
    "id", "secao", "sub_secao_ambiente", "item", "descricao_detalhada",
    "quantidade", "unidade",
    "medida_original", "largura", "altura", "profundidade",
    "material", "acabamento_cor", "tipo_producao",
    "status_arte", "status_compra_locacao",
    "observacoes_tecnicas", "pendencia_acao_necessaria", "nivel_atencao",
]
FORN_WIDTHS = {
    1:10, 2:22, 3:24, 4:38, 5:44,
    6:10, 7:10,
    8:22, 9:10, 10:10, 11:13,
    12:22, 13:22, 14:22,
    15:22, 16:26,
    17:38, 18:38, 19:13,
}
_CENTER_FORN = {1, 6, 7, 9, 10, 11, 19}

# Rotating palette for supplier tab colours
_TAB_COLORS = [
    "4472C4", "ED7D31", "A9D18E", "FF0000", "FFC000",
    "5B9BD5", "70AD47", "FF7070", "C55A11", "7030A0",
    "00B050", "9E480E", "833C00", "375623", "843C0C",
]


def _sanitize_tab(name: str) -> str:
    """Return a valid Excel sheet-tab name (max 31 chars, no special chars)."""
    cleaned = re.sub(r'[\/\\\[\]\?\*:]', '-', name)
    return cleaned[:31].strip()


def _build_fornecedor_sheet(ws, fornecedor: str, itens: list[dict], tab_color: str):
    ws.sheet_view.showGridLines = False

    # Title
    n_cols = len(FORN_HEADERS)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = f"FORNECEDOR: {fornecedor.upper()}"
    t.font  = Font(name="Calibri", bold=True, size=12, color=_WHITE)
    t.fill  = _solid(tab_color)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 28)

    # Sub-header counts
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    sub = ws["A2"]
    sub.value = f"{len(itens)} item(s) · Fornecedor: {fornecedor}"
    sub.font  = Font(name="Calibri", size=9, color="444444")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    sub.fill  = _solid(_GRAY)
    _row_height(ws, 2, 16)

    # Headers row 3
    for col, label in enumerate(FORN_HEADERS, start=1):
        _hdr_cell(ws, 3, col, label, bg=tab_color, size=9)
    _row_height(ws, 3, 24)

    _freeze(ws, "A4")
    _set_widths(ws, FORN_WIDTHS)

    for r_idx, item in enumerate(itens, start=4):
        nivel    = str(item.get("nivel_atencao", "")).strip().lower()
        nivel_bg = NIVEL_FILL.get(nivel)
        alt_bg   = _ALT if r_idx % 2 == 0 else None

        for c_idx, field in enumerate(FORN_FIELDS, start=1):
            val   = item.get(field, "") or ""
            h_al  = "center" if c_idx in _CENTER_FORN else "left"
            _data_cell(ws, r_idx, c_idx, val, bg=alt_bg, h=h_al)

        if nivel_bg:
            cell = ws.cell(row=r_idx, column=19)
            cell.fill = _solid(nivel_bg)
            cell.font = Font(name="Calibri", bold=True, size=9)

        _row_height(ws, r_idx, 32)

    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}{3 + len(itens)}"
    ws.sheet_properties.tabColor = tab_color


# ─────────────────────────────────────────────────────────────────────────────
# Main builder
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(data: dict, filename: str) -> bytes:
    """
    Build the full Excel workbook from analyzed data and return as bytes.

    data:     output dict from local_analyzer.analyze_document()
    filename: original uploaded filename (displayed in Resumo)
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    itens     = data.get("itens_detalhados", [])
    pendencias= data.get("pendencias", [])

    # Lookup dict: item id → item
    item_lookup = {it.get("id", ""): it for it in itens}

    # ── Fixed sheets ──────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Resumo");            _build_resumo(ws1, data, filename)
    ws2 = wb.create_sheet("Itens Detalhados");  _build_itens(ws2, itens)
    ws3 = wb.create_sheet("Por Seção");         _build_por_secao(ws3, itens)
    ws4 = wb.create_sheet("Medições");          _build_medicoes(ws4, itens)
    ws5 = wb.create_sheet("Pendências");        _build_pendencias(ws5, pendencias, item_lookup)
    ws6 = wb.create_sheet("Legenda e Status");  _build_legenda(ws6, data.get("legenda_status", []))

    # Tab colours for fixed sheets
    ws1.sheet_properties.tabColor = _NAV
    ws2.sheet_properties.tabColor = "2E75B6"
    ws3.sheet_properties.tabColor = "4472C4"
    ws4.sheet_properties.tabColor = "70AD47"
    ws5.sheet_properties.tabColor = "FF6B6B"
    ws6.sheet_properties.tabColor = "FFC000"

    # ── Supplier sheets (at the end) ──────────────────────────────────────────
    # Group items by supplier, skip "Não informado"
    suppliers: dict[str, list[dict]] = {}
    no_supplier: list[dict] = []

    for item in itens:
        forn = (item.get("fornecedor") or "").strip()
        if not forn or forn.lower() in ("não informado", "nao informado", "a confirmar", ""):
            no_supplier.append(item)
        else:
            suppliers.setdefault(forn, []).append(item)

    # Sort suppliers alphabetically
    for forn_idx, (forn_name, forn_items) in enumerate(sorted(suppliers.items())):
        tab_color = _TAB_COLORS[forn_idx % len(_TAB_COLORS)]
        tab_name  = _sanitize_tab(forn_name)
        ws_forn   = wb.create_sheet(tab_name)
        _build_fornecedor_sheet(ws_forn, forn_name, forn_items, tab_color)

    # "Sem Fornecedor" tab if there are unassigned items
    if no_supplier:
        ws_no = wb.create_sheet("Sem Fornecedor")
        _build_fornecedor_sheet(ws_no, "Sem Fornecedor / A Definir", no_supplier, "808080")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
